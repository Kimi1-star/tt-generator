"""
TT (Commercial Invoice / Packing List / Shipment Advice) Excel generator.
Layout: A4 portrait, horizontally centred, no gridlines, 1-page-wide fit.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

# ── Fixed constants ─────────────────────────────────────────────────────────
COMPANY_NAME = "SHARPMAX INTERNATIONAL (HONG KONG) CO., LIMITED"
COMPANY_ADDR = ("ROOM 704, 7/F., TOWER A, NEW MANDARIN PLAZA, "
                "14 SCIENCE MUSEUM ROAD, TST EAST, KOWLOON, HONG KONG")
BENEFICIARY  = "SHARPMAX INTERNATIONAL (HONGKONG) CO.,LIMITED"
BANK_NAME    = "CHINA ZHESHANG BANK"
BANK_ADDR    = "NO.288 QINGCHUN ROAD HANG ZHOU,ZHEJIANG,CHINA"
SWIFT_CODE   = "ZJCBCN2N"
ACCOUNT_NO   = "NRA3310010711420100000726"

# ── Number → English words ──────────────────────────────────────────────────
_ONES = ['', 'ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT',
         'NINE', 'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN',
         'SIXTEEN', 'SEVENTEEN', 'EIGHTEEN', 'NINETEEN']
_TENS = ['', '', 'TWENTY', 'THIRTY', 'FORTY', 'FIFTY',
         'SIXTY', 'SEVENTY', 'EIGHTY', 'NINETY']

def _below_thousand(n):
    if n == 0: return ''
    if n < 20:  return _ONES[n]
    if n < 100: return _TENS[n//10] + (' ' + _ONES[n%10] if n%10 else '')
    rem = n % 100
    return _ONES[n//100] + ' HUNDRED' + (' AND ' + _below_thousand(rem) if rem else '')

def amount_to_words(amount):
    dollars, cents = int(amount), round((amount - int(amount)) * 100)
    def _full(n):
        if n == 0: return 'ZERO'
        parts = []
        if n >= 1_000_000: parts.append(_below_thousand(n//1_000_000) + ' MILLION');  n %= 1_000_000
        if n >= 1_000:     parts.append(_below_thousand(n//1_000)     + ' THOUSAND'); n %= 1_000
        if n > 0: parts.append(_below_thousand(n))
        return ' '.join(parts)
    txt = 'SAY US DOLLARS ' + _full(dollars)
    if cents: txt += ' AND CENTS ' + _below_thousand(cents)
    return txt + ' ONLY'


# ── Style helpers ────────────────────────────────────────────────────────────
def _thin():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, size=9):
    return Font(name='Arial', bold=bold, size=size)

def _align(h='left', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _c(ws, row, col, value='', bold=False, size=9,
       h='left', v='center', wrap=True, border=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = _font(bold, size)
    c.alignment = _align(h, v, wrap)
    if border: c.border = _thin()
    if fmt:    c.number_format = fmt
    return c

def _mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

def _rh(ws, row, height):
    ws.row_dimensions[row].height = height


# ── A4 page setup ────────────────────────────────────────────────────────────
def _setup_a4(ws, last_row, ncols):
    col_letter = 'ABCDE'[ncols - 1]
    ws.print_area                      = f'A1:{col_letter}{last_row}'
    ws.page_setup.paperSize            = 9          # A4
    ws.page_setup.orientation          = 'portrait'
    ws.page_setup.fitToPage            = True
    ws.page_setup.fitToWidth           = 1
    ws.page_setup.fitToHeight          = 99
    ws.page_margins                    = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True      # centre on page
    ws.sheet_view.showGridLines         = False     # no dashed lines


# ── CI Sheet ─────────────────────────────────────────────────────────────────
# 5 columns  A=GRADE/labels  B=SIZE  C=QUANTITY  D=UNIT PRICE  E=AMOUNT
# Widths: A=26  B=12  C=11  D=11  E=16  → total 76 units ≈ 174 mm
# Horizontally centred → equal margins on A4 (≈ 18 mm each side)
def _build_ci(ws, d):
    for col, w in zip('ABCDE', [26, 12, 11, 11, 16]):
        ws.column_dimensions[col].width = w

    r = 1

    # ── Company header
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_NAME, bold=True, size=11, h='center', wrap=False)
    _rh(ws, r, 18); r += 1
    # Address: allow wrap so long text shows fully
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_ADDR, size=9, h='center', wrap=True)
    _rh(ws, r, 26); r += 1
    r += 1  # blank

    # ── Title
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, 'COMMERCIAL INVOICE', bold=True, size=14, h='center', wrap=False)
    _rh(ws, r, 22); r += 1
    r += 2  # blank rows

    # ── TO: + ORIGINAL on same row  (ORIGINAL spans D:E)
    _c(ws, r, 1, 'TO:', bold=True)
    _mg(ws, r, 4, r, 5); _c(ws, r, 4, 'ORIGINAL', bold=True, h='right', wrap=False)
    r += 1

    # ── Buyer (A:C) | DATE + INVOICE NO. merged D:E, top-right
    buyer = [d.get('buyer_name', '')]
    if d.get('buyer_address'): buyer.append('ADD: ' + d['buyer_address'])
    if d.get('buyer_ref'):     buyer.append(d['buyer_ref'])
    _mg(ws, r, 1, r, 3); _c(ws, r, 1, '\n'.join(buyer))
    date_inv = f"DATE: {d['invoice_date']}\nINVOICE NO.: {d['invoice_no']}"
    _mg(ws, r, 4, r, 5); _c(ws, r, 4, date_inv, h='right', v='top', wrap=True)
    _rh(ws, r, 48); r += 1

    r += 2  # blank rows

    # ── Goods description
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, 'DESCRIPTION OF GOODS & SERVICES', bold=True); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"COMMODITY: {d['commodity']}"); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"SALES CONTRACT NO.: {d['contract_no']}"); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"COUNTRY OF ORIGIN: {d.get('country_of_origin','CHINA')}"); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"PRICE TERMS: {d['price_terms']}"); r += 1
    r += 1  # blank

    # ── Table headers (2 rows)
    # A=GRADE  B=SIZE  C=QUANTITY  D=UNIT PRICE  E=AMOUNT
    for col, (h1, h2) in enumerate(zip(
            ['GRADE', 'SIZE',  'QUANTITY',  'UNIT PRICE', 'AMOUNT'],
            ['',      '(MM)',  '(MT)',       '(USD/MT)',   '(USD)']), 1):
        _c(ws, r,   col, h1, bold=True, h='center', border=True)
        _c(ws, r+1, col, h2, bold=True, h='center', border=True)
    r += 2

    # ── Data rows
    total_qty = total_amt = 0.0
    for g in d['grades']:
        qty   = round(float(g['quantity_mt']), 3)
        price = round(float(g['unit_price']),  2)
        amt   = round(qty * price, 2)
        total_qty = round(total_qty + qty, 3)
        total_amt = round(total_amt + amt, 2)
        for col, (v, f) in enumerate(zip(
                [g['grade'], str(g['size']), qty, price, amt],
                [None, None, '#,##0.000', '#,##0.00', '#,##0.00']), 1):
            _c(ws, r, col, v, h='right' if col >= 3 else 'center', border=True, fmt=f)
        r += 1

    # ── TOTAL row — TOTAL label in col A (same column as GRADE)
    for col, (v, f) in enumerate(zip(
            ['TOTAL', '', total_qty, '', total_amt],
            [None, None, '#,##0.000', None, '#,##0.00']), 1):
        _c(ws, r, col, v, bold=True,
           h='right' if col in (3, 5) else 'center', border=True, fmt=f)
    r += 1

    # ── Prepayment / balance
    fob     = round(total_amt - float(d.get('freight') or 0), 2)
    freight = float(d.get('freight') or 0)

    if d.get('has_prepayment') and d.get('prepayment'):
        prepayment = round(float(d['prepayment']), 2)
        balance    = round(total_amt - prepayment, 2)
        _c(ws, r, 1, 'DEDUCTION (DOWN PAYMENT)', bold=True)
        _c(ws, r, 5, prepayment, h='right', fmt='#,##0.00'); r += 1
        _c(ws, r, 1, 'BALANCE', bold=True)
        _c(ws, r, 5, balance, bold=True, h='right', fmt='#,##0.00'); r += 1
        say_amount = balance
    else:
        say_amount = total_amt

    # ── SAY row
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, amount_to_words(say_amount))
    _rh(ws, r, 28); r += 1

    # ── Tail (differs by payment type)
    if d.get('has_prepayment') and d.get('prepayment'):
        r += 1
        _mg(ws, r, 1, r, 5); _c(ws, r, 1, f'BENEFICIARY: {BENEFICIARY}'); r += 1
        _mg(ws, r, 1, r, 5); _c(ws, r, 1, f'ADD: {COMPANY_ADDR}'); r += 1
        for lbl, val in [('BANK NAME', BANK_NAME), ('BANK ADDRESS', BANK_ADDR),
                         ('SWIFT CODE', SWIFT_CODE), ('ACCOUNT NO.', ACCOUNT_NO)]:
            _mg(ws, r, 1, r, 5)
            _c(ws, r, 1, f'{lbl}:{val}' if lbl == 'SWIFT CODE' else f'{lbl}: {val}'); r += 1
        _c(ws, r, 1, 'FOB VALUE: USD')
        _c(ws, r, 3, fob, h='right', fmt='#,##0.00'); r += 1
        _c(ws, r, 1, 'FREIGHT AMOUNT: USD')
        _c(ws, r, 3, freight, h='right', fmt='#,##0.00'); r += 1
    else:
        pay_txt = ("THE INVOICE AMOUNT SHOULD BE PAID WITHIN 5 WORKING DAYS AGAINST "
                   "BL SCANNED COPY, COMMERCIAL INVOICE, PACKING LIST AT THE FIRST "
                   "VERSION SENT BY THE SELLER THROUGH THE EMAIL.")
        _mg(ws, r, 1, r, 5); _c(ws, r, 1, pay_txt)
        _rh(ws, r, 35); r += 1
        _c(ws, r, 1, 'FOB VALUE: USD')
        _c(ws, r, 2, fob, h='right', fmt='#,##0.00'); r += 1
        _c(ws, r, 1, 'FREIGHT AMOUNT: USD')
        _c(ws, r, 2, freight, h='right', fmt='#,##0.00'); r += 2
        _mg(ws, r, 1, r, 5); _c(ws, r, 1, f'BENEFICIARY: {BENEFICIARY}'); r += 1
        _mg(ws, r, 1, r, 5); _c(ws, r, 1, f'ADD: {COMPANY_ADDR}'); r += 1
        for lbl, val in [('BANK NAME', BANK_NAME), ('BANK ADDRESS', BANK_ADDR),
                         ('SWIFT CODE', SWIFT_CODE), ('ACCOUNT NO.', ACCOUNT_NO)]:
            _mg(ws, r, 1, r, 5)
            _c(ws, r, 1, f'{lbl}:{val}' if lbl == 'SWIFT CODE' else f'{lbl}: {val}'); r += 1

    r += 2
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_NAME, bold=True, h='center', wrap=False)

    _setup_a4(ws, r, ncols=5)


# ── PL Sheet ─────────────────────────────────────────────────────────────────
# 5 columns  A=GRADE  B=SIZE  C=NUMBER OF COILS  D=NET WEIGHT  E=GROSS WEIGHT
# Widths: A=22  B=12  C=13  D=14  E=14  → total 75 units ≈ 172 mm
def _build_pl(ws, d):
    for col, w in zip('ABCDE', [22, 12, 13, 14, 14]):
        ws.column_dimensions[col].width = w

    r = 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_NAME, bold=True, size=11, h='center', wrap=False)
    _rh(ws, r, 18); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_ADDR, size=9, h='center', wrap=True)
    _rh(ws, r, 26); r += 1
    r += 1

    _mg(ws, r, 1, r, 5); _c(ws, r, 1, 'PACKING LIST', bold=True, size=14, h='center', wrap=False)
    _rh(ws, r, 22); r += 1
    r += 2

    _c(ws, r, 1, 'TO:', bold=True)
    _mg(ws, r, 4, r, 5); _c(ws, r, 4, 'ORIGINAL', bold=True, h='right', wrap=False)
    r += 1

    buyer = [d.get('buyer_name', '')]
    if d.get('buyer_address'): buyer.append('ADD: ' + d['buyer_address'])
    if d.get('buyer_ref'):     buyer.append(d['buyer_ref'])
    _mg(ws, r, 1, r, 3); _c(ws, r, 1, '\n'.join(buyer))
    date_inv = f"DATE: {d['invoice_date']}\nINVOICE NO.: {d['invoice_no']}"
    _mg(ws, r, 4, r, 5); _c(ws, r, 4, date_inv, h='right', v='top', wrap=True)
    _rh(ws, r, 48); r += 1

    r += 2

    _mg(ws, r, 1, r, 5); _c(ws, r, 1, 'DESCRIPTION OF GOODS & SERVICES', bold=True); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"COMMODITY: {d['commodity']}"); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"SALES CONTRACT NO.: {d['contract_no']}"); r += 1
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, f"COUNTRY OF ORIGIN: {d.get('country_of_origin','CHINA')}"); r += 1
    r += 1

    # Table header
    for col, (h1, h2) in enumerate(zip(
            ['GRADE', 'SIZE',  'NUMBER OF', 'NET WEIGHT /\nQUANTITY', 'GROSS WEIGHT'],
            ['',      '(MM)',  'COILS',     '(MT)',                   '(MT)']), 1):
        _c(ws, r,   col, h1, bold=True, h='center', border=True)
        _c(ws, r+1, col, h2, bold=True, h='center', border=True)
    r += 2

    total_coils = 0; total_net = total_gross = 0.0
    for g in d['grades']:
        coils = int(g.get('num_coils', 0))
        net   = round(float(g['quantity_mt']), 3)
        gross = round(float(g.get('gross_weight_mt') or net), 3)
        total_coils += coils
        total_net    = round(total_net   + net,   3)
        total_gross  = round(total_gross + gross, 3)
        for col, (v, f) in enumerate(zip(
                [g['grade'], str(g['size']), coils, net, gross],
                [None, None, '#,##0', '#,##0.000', '#,##0.000']), 1):
            _c(ws, r, col, v, h='right' if col >= 3 else 'center', border=True, fmt=f)
        r += 1

    for col, (v, f) in enumerate(zip(
            ['TOTAL', '', total_coils, total_net, total_gross],
            [None, None, '#,##0', '#,##0.000', '#,##0.000']), 1):
        _c(ws, r, col, v, bold=True,
           h='right' if col >= 3 else 'center', border=True, fmt=f)
    r += 1

    r += 3
    _mg(ws, r, 1, r, 5); _c(ws, r, 1, COMPANY_NAME, bold=True, h='center', wrap=False)
    _setup_a4(ws, r, ncols=5)


# ── SA Sheet ─────────────────────────────────────────────────────────────────
# 4 columns  A=label  B=value  C-D=empty/merged
# Widths: A=32  B=18  C=14  D=12  → total 76 units ≈ 174 mm
def _build_sa(ws, d):
    for col, w in zip('ABCD', [32, 18, 14, 12]):
        ws.column_dimensions[col].width = w

    r = 1
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, COMPANY_NAME, bold=True, size=11, h='center', wrap=False)
    _rh(ws, r, 18); r += 1
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, COMPANY_ADDR, size=9, h='center', wrap=True)
    _rh(ws, r, 26); r += 1
    r += 1

    _mg(ws, r, 1, r, 4); _c(ws, r, 1, 'SHIPMENT ADVICE', bold=True, size=14, h='center', wrap=False)
    _rh(ws, r, 22); r += 1
    r += 2

    _c(ws, r, 1, 'TO:', bold=True)
    _c(ws, r, 4, 'ORIGINAL', bold=True, h='right', v='top', wrap=False)
    r += 1

    buyer = [d.get('buyer_name', '')]
    if d.get('buyer_address'): buyer.append('ADD: ' + d['buyer_address'])
    if d.get('buyer_ref'):     buyer.append(d['buyer_ref'])
    _mg(ws, r, 1, r, 3); _c(ws, r, 1, '\n'.join(buyer))
    date_inv = f"DATE: {d['invoice_date']}\nINVOICE NO.: {d['invoice_no']}"
    _c(ws, r, 4, date_inv, h='right', v='top', wrap=True)
    _rh(ws, r, 48); r += 1

    r += 2

    _mg(ws, r, 1, r, 4); _c(ws, r, 1, 'SHIPMENT DETAILS', bold=True); r += 1
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, f"COMMODITY: {d['commodity']}"); r += 1

    total_qty = round(sum(float(g['quantity_mt']) for g in d['grades']), 3)
    total_amt = round(sum(float(g['quantity_mt']) * float(g['unit_price'])
                         for g in d['grades']), 2)

    _c(ws, r, 1, 'QUANTITY LOADED: MT')
    _c(ws, r, 2, total_qty, h='right', fmt='#,##0.000'); r += 1
    _c(ws, r, 1, 'INVOICE VALUE: USD')
    _c(ws, r, 2, total_amt, h='right', fmt='#,##0.00'); r += 1
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, f"NAME OF VESSEL: {d.get('vessel','')}"); r += 1
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, f"PORT OF LOADING: {d.get('loading_port','')}"); r += 1
    eta = d.get('eta_etd_label', 'ETA').upper()
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, f"{eta}: {d.get('eta_etd_date','')}"); r += 1

    r += 15
    _mg(ws, r, 1, r, 4); _c(ws, r, 1, COMPANY_NAME, bold=True, h='center', wrap=False)
    _setup_a4(ws, r, ncols=4)


# ── Entry point ──────────────────────────────────────────────────────────────
def generate_tt_excel(data):
    wb = Workbook()
    wb.remove(wb.active)
    _build_ci(wb.create_sheet('CI'),  data)
    _build_pl(wb.create_sheet('PL'),  data)
    _build_sa(wb.create_sheet('SA'),  data)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
